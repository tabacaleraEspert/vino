"""Service for extracting transactions from bank statements using OpenAI."""
from __future__ import annotations

import base64
import hashlib
import json
import logging
import re
from typing import Any

from openai import AsyncOpenAI

from app.core.config import settings

logger = logging.getLogger(__name__)

_client: AsyncOpenAI | None = None


def _get_client() -> AsyncOpenAI:
    global _client
    if _client is None:
        _client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
    return _client


_SYSTEM_PROMPT = """Sos un extractor de transacciones financieras de resumenes bancarios y de tarjetas de credito argentinas.

Extrae TODAS las transacciones/movimientos del documento. Para cada una devolvé:
- fecha: en formato YYYY-MM-DD. Si solo hay mes/año, usá el día 1.
- descripcion: nombre del comercio o descripción tal como aparece, limpio y legible.
- monto: número positivo (sin signo, sin $ ni separadores de miles). Decimales con punto.
- moneda: "ARS", "USD", etc. Si no se indica, asumir ARS.
- tipo: "Gasto" para compras/débitos, "Ingreso" para créditos/pagos a favor.

Las categorías disponibles del usuario son:
{categories_json}

Para cada transacción, sugerí:
- categoria_id: el ID de la categoría que mejor aplica (o null si no estás seguro)
- subcategoria_id: el ID de la subcategoría que mejor aplica (o null)

NO incluyas:
- Totales, subtotales, saldos anteriores, impuestos globales, intereses del resumen
- Filas que sean encabezados o resúmenes

Respondé SOLO con JSON válido:
{
  "transactions": [
    {
      "fecha": "2026-04-15",
      "descripcion": "MERCADOLIBRE",
      "monto": 15000.50,
      "moneda": "ARS",
      "tipo": "Gasto",
      "categoria_id": 3,
      "subcategoria_id": 12
    }
  ],
  "statement_period": "2026-04",
  "card_or_account": "Visa terminada en 1234"
}"""


def _build_categories_json(categories: list[dict[str, Any]]) -> str:
    """Build a compact JSON of categories for the AI prompt."""
    cats = []
    for c in categories:
        entry: dict[str, Any] = {"id": c.get("id"), "nombre": c.get("nombre", "")}
        subs = c.get("subcategorias", [])
        if subs:
            entry["subcategorias"] = [
                {"id": s.get("id"), "nombre": s.get("nombre", "")} for s in subs
            ]
        cats.append(entry)
    return json.dumps(cats, ensure_ascii=False)


def make_origen_id(fecha: str, monto: float, descripcion: str) -> str:
    """Generate a deterministic origin ID for deduplication."""
    raw = f"{fecha}_{monto}_{descripcion}".lower().strip()
    return f"stmt_{hashlib.md5(raw.encode()).hexdigest()[:12]}"


async def extract_from_text(text: str, categories: list[dict]) -> dict:
    """Extract transactions from plain text (PDF text extraction)."""
    client = _get_client()
    prompt = _SYSTEM_PROMPT.replace("{categories_json}", _build_categories_json(categories))

    response = await client.chat.completions.create(
        model="gpt-4.1-mini",
        messages=[
            {"role": "system", "content": prompt},
            {"role": "user", "content": f"Extracto bancario:\n\n{text}"},
        ],
        response_format={"type": "json_object"},
        max_completion_tokens=4000,
    )

    raw = response.choices[0].message.content or "{}"
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        logger.error("Failed to parse OpenAI response as JSON: %s", raw[:200])
        return {"transactions": [], "error": "No se pudo parsear la respuesta"}


async def extract_from_image(image_bytes: bytes, mime_type: str, categories: list[dict]) -> dict:
    """Extract transactions from an image using Vision API."""
    client = _get_client()
    prompt = _SYSTEM_PROMPT.replace("{categories_json}", _build_categories_json(categories))
    b64 = base64.b64encode(image_bytes).decode("utf-8")

    response = await client.chat.completions.create(
        model="gpt-4.1-mini",
        messages=[
            {"role": "system", "content": prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "Extraé las transacciones de este extracto bancario:"},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:{mime_type};base64,{b64}", "detail": "high"},
                    },
                ],
            },
        ],
        response_format={"type": "json_object"},
        max_completion_tokens=4000,
    )

    raw = response.choices[0].message.content or "{}"
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        logger.error("Failed to parse Vision response as JSON: %s", raw[:200])
        return {"transactions": [], "error": "No se pudo parsear la respuesta"}


_MONTH_MAP = {
    "ene": "01", "enero": "01", "feb": "02", "febrero": "02",
    "mar": "03", "marzo": "03", "abr": "04", "abril": "04",
    "may": "05", "mayo": "05", "jun": "06", "junio": "06",
    "jul": "07", "julio": "07", "ago": "08", "agosto": "08",
    "set": "09", "setiem": "09", "sep": "09", "septiem": "09",
    "oct": "10", "octubre": "10", "nov": "11", "noviem": "11",
    "dic": "12", "diciem": "12",
}

# Regex for Santander PDF transaction lines
# Examples:
#   25 Octubre 02 061353 *  NIKE ALCORTA                C.04/06     17.499,83
#   06 888689    OPENAI *CHATGPT  in1SfLrfCUSD       20,00                    20,00
#   02 000071 *  THE KOOPLES                 C.04/06     30.500,00
_SANTANDER_TX_RE = re.compile(
    r"^\s*(?:(\d{1,2})\s+([A-Za-záéíóú]+)\.?\s+)?"  # optional: day + month name
    r"(\d{1,2})\s+"                                    # day of transaction
    r"(\d{5,6})\s*[*K ]?\s*"                           # comprobante number + type
    r"(.+?)"                                           # description (greedy but trimmed)
    r"(?:\s+C\.(\d{2}/\d{2}))?"                        # optional cuotas C.XX/XX
    r"\s+"                                             # separator
    r"([\d.,]+(?:-)?)"                                 # amount (may end with -)
    r"\s*$"
)


def _parse_santander_pdf_text(full_text: str) -> dict:
    """Parse Santander Visa PDF text format into transactions."""
    lines = full_text.split("\n")

    transactions = []
    current_month_str = ""  # "01" for Enero
    current_year = ""       # "26" or "2026"
    statement_period = ""
    card_info = "Visa Santander"
    last_date = ""

    # Skip patterns
    skip_patterns = [
        "SALDO ANTERIOR", "SU PAGO EN PESOS", "TRANSFERENCIA DEUDA",
        "IMPUESTO DE SELLOS", "INTERESES FINANC", "DB IVA", "IIBB PERCEP",
        "IVA RG 4240", "DB.RG 5617", "Tarjeta", "Total Consumos",
        "Plan V:", "fijas su saldo", "cuotas de $", "TNA:", "CFTEA",
    ]

    # Extract CIERRE date and VENCIMIENTO for year reference
    for line in lines:
        # "CIERRE  29 Ene 26 VENCIMIENTO 06 Feb 26"
        m = re.search(r"CIERRE\s+(\d{1,2})\s+(\w+)\s+(\d{2,4})", line)
        if m:
            day, mon, yr = m.group(1), m.group(2).lower(), m.group(3)
            for key, val in _MONTH_MAP.items():
                if mon.startswith(key):
                    current_year = yr if len(yr) == 4 else f"20{yr}"
                    statement_period = f"{current_year}-{val}"
                    break
            break
        # Also try "VENCIMIENTO DD Mes YY" format
        m = re.search(r"VENCIMIENTO\s+(\d{1,2})\s+(\w+)\s+(\d{2,4})", line)
        if m:
            day, mon, yr = m.group(1), m.group(2).lower(), m.group(3)
            for key, val in _MONTH_MAP.items():
                if mon.startswith(key):
                    current_year = yr if len(yr) == 4 else f"20{yr}"
                    statement_period = f"{current_year}-{val}"
                    break
            break

    if not current_year:
        # Try extracting from filename or default
        current_year = "2026"

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Skip boilerplate and non-transaction lines
        if any(skip in line for skip in skip_patterns):
            continue

        # Detect month headers like "25 Octubre", "26 Enero", "25 Setiem."
        month_header = re.match(r"^(\d{1,2})\s+([A-Za-záéíóú]+)\.?\s+(\d{1,2})\s", line)
        if month_header:
            month_name = month_header.group(2).lower()
            for key, val in _MONTH_MAP.items():
                if month_name.startswith(key):
                    current_month_str = val
                    break
            # Adjust year: if month > current statement month, it's previous year
            if statement_period and current_month_str:
                stmt_month = int(statement_period.split("-")[1])
                tx_month = int(current_month_str)
                if tx_month > stmt_month:
                    tx_year = str(int(current_year) - 1)
                else:
                    tx_year = current_year
            else:
                tx_year = current_year

        # Try to parse as transaction
        # Simplified: look for comprobante pattern (5-6 digits followed by * or K)
        tx_match = re.match(
            r"^\s*(?:\d{1,2}\s+[A-Za-záéíóú]+\.?\s+)?"  # optional month header
            r"(\d{1,2})\s+"                               # day
            r"(\d{5,6})\s*([*K ]?)\s*"                    # comprobante + type
            r"(.+?)$",                                     # rest of line
            line
        )

        if not tx_match:
            continue

        day = tx_match.group(1).zfill(2)
        desc_and_amount = tx_match.group(4).strip()

        if not current_month_str:
            continue

        # Determine year for this transaction
        if statement_period:
            stmt_month = int(statement_period.split("-")[1])
            tx_month = int(current_month_str)
            tx_year = str(int(current_year) - 1) if tx_month > stmt_month else current_year
        else:
            tx_year = current_year

        fecha = f"{tx_year}-{current_month_str}-{day}"

        # Parse description and amounts from the rest
        # The amounts are at the end, separated by spaces
        # Pattern: DESCRIPTION [C.XX/XX] [USD_AMOUNT] PESOS_AMOUNT [USD_AMOUNT]

        # Try to extract cuotas
        cuotas = ""
        cuotas_match = re.search(r"C\.(\d{2}/\d{2,3})", desc_and_amount)
        if cuotas_match:
            cuotas = cuotas_match.group(1)

        # Extract amounts (numbers with dots and commas at end of line)
        # Amounts can be: 17.499,83 or 20,00 or 2,99 or 133.309,66
        amounts = re.findall(r"(\d[\d.]*,\d{2}-?)", desc_and_amount)

        if not amounts:
            continue

        # Clean description: remove amounts, cuotas, USD markers, reference codes
        desc_clean = desc_and_amount
        for amt in amounts:
            desc_clean = desc_clean.replace(amt, "")
        desc_clean = re.sub(r"C\.\d{2}/\d{2,3}", "", desc_clean)
        desc_clean = re.sub(r"\bUSD\b", "", desc_clean)
        # Remove long alphanumeric reference codes (e.g., in1SfLrfC, MTZ476WY1, hLhTMW5nA)
        # but keep actual names — only remove if it looks like a hash (mixed case + digits)
        desc_clean = re.sub(r"\b[a-zA-Z0-9]*\d+[a-zA-Z]+[a-zA-Z0-9]*\b", "", desc_clean)
        # Remove long number sequences like 000000041700043, 300070577571001
        desc_clean = re.sub(r"\b\d{10,}\b", "", desc_clean)
        desc_clean = re.sub(r"\s+", " ", desc_clean).strip()

        if not desc_clean or len(desc_clean) < 2:
            continue

        if cuotas:
            desc_clean = f"{desc_clean} ({cuotas})"

        # Determine if USD or ARS
        has_usd = "USD" in desc_and_amount

        # Parse the main amount
        if has_usd and len(amounts) >= 1:
            # USD transaction - typically the last amount is USD
            amount_str = amounts[-1]
            monto = _parse_santander_amount(amount_str)
            moneda = "USD"
        else:
            # ARS transaction
            amount_str = amounts[-1]
            monto = _parse_santander_amount(amount_str)
            moneda = "ARS"

        if monto is None or monto == 0:
            continue

        is_negative = amount_str.endswith("-")
        tipo = "Ingreso" if is_negative else "Gasto"

        transactions.append({
            "fecha": fecha,
            "descripcion": desc_clean,
            "monto": round(abs(monto), 2),
            "moneda": moneda,
            "tipo": tipo,
            "categoria_id": None,
            "subcategoria_id": None,
        })

    return {
        "transactions": transactions,
        "statement_period": statement_period,
        "card_or_account": card_info,
    }


def _parse_santander_amount(s: str) -> float | None:
    """Parse Santander amount: 17.499,83 or 2,99 or 133.309,66-"""
    s = s.strip().rstrip("-")
    if not s:
        return None
    # Santander uses dots for thousands, comma for decimals
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_bbva_pdf_text(full_text: str) -> dict:
    """Parse BBVA PDF — auto-detects credit card vs caja de ahorro format."""
    # Credit card format has "Tarjetas de Crédito" or "Consumos"
    is_credit_card = "Tarjetas de Cr" in full_text or "Mastercard" in full_text or "VISA" in full_text
    if is_credit_card and "Consumos" in full_text:
        return _parse_bbva_credit_card(full_text)
    return _parse_bbva_caja_ahorro(full_text)


_BBVA_MONTH_MAP = {
    "ene": "01", "feb": "02", "mar": "03", "abr": "04",
    "may": "05", "jun": "06", "jul": "07", "ago": "08",
    "sep": "09", "oct": "10", "nov": "11", "dic": "12",
}


def _parse_bbva_credit_card(full_text: str) -> dict:
    """Parse BBVA credit card (Mastercard/Visa) PDF."""
    lines = full_text.split("\n")
    transactions: list[dict] = []

    # Detect card type
    card_info = "BBVA Tarjeta de Crédito"
    if "Mastercard" in full_text:
        card_info = "BBVA Mastercard"
    elif "VISA" in full_text:
        card_info = "BBVA Visa"

    # Skip patterns
    skip_patterns = [
        "SU PAGO", "CR.RG", "DB.RG", "TOTAL CONSUMOS", "SALDO ACTUAL",
        "SALDO ANTERIOR", "DEVOLUCION", "IIBB PERCEP", "IVA RG",
        "IMPUESTO", "COMISION", "COM CUENTA",
    ]

    # Parse sections: only extract from "Consumos ..." sections, skip "Sus pagos" and "Impuestos"
    in_consumos = False
    stop_sections = [
        "Impuestos, cargos e intereses", "Legales y avisos",
        "Sus pagos y ajustes", "SALDO ACTUAL", "TOTAL CONSUMOS",
    ]

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Detect "Consumos [Name]" section
        if re.match(r"^Consumos\s+", stripped):
            in_consumos = True
            continue

        # Detect stop sections
        if any(s in stripped for s in stop_sections):
            in_consumos = False
            continue

        # Skip headers
        if "FECHA" in stripped and "DESCRIPCIÓN" in stripped:
            continue

        if not in_consumos:
            continue

        # Skip non-transaction rows
        if any(skip in stripped for skip in skip_patterns):
            continue

        # Parse: DD-Mon-YY  DESCRIPTION  [NRO.CUPÓN]  [PESOS]  [DÓLARES]
        # Examples:
        #   04-Mar-26 PERSONAL PERSFLOW75010002 03/2 043783 107.726,80
        #   06-Mar-26 OPENAI *CHATGPT SUBSCR USD 20,00 599549 20,00
        #   16-Mar-26 LinkedInPreA *23461733 USD 1,44 211122 1,44
        m = re.match(r"^\s*(\d{2})-(\w{3})-(\d{2})\s+(.+)$", stripped)
        if not m:
            continue

        day = m.group(1)
        month_str = m.group(2).lower()
        year = f"20{m.group(3)}"
        month = _BBVA_MONTH_MAP.get(month_str[:3])
        if not month:
            continue

        rest = m.group(4).strip()

        # Extract amounts from end
        amount_pattern = r"-?[\d]+(?:\.[\d]{3})*,\d{2}"
        amounts = re.findall(amount_pattern, rest)

        if not amounts:
            continue

        # Remove amounts and cupón number from description
        desc = rest
        for amt in amounts:
            desc = desc.replace(amt, "", 1)
        # Remove 6-digit cupón numbers
        desc = re.sub(r"\b\d{6}\b", "", desc)
        # Remove USD marker
        has_usd = "USD" in desc
        desc = re.sub(r"\bUSD\b", "", desc)
        desc = re.sub(r"\s+", " ", desc).strip()

        if not desc or len(desc) < 2:
            continue

        # Determine currency and amount
        if has_usd and amounts:
            monto = _parse_ar_money(amounts[-1])
            moneda = "USD"
        else:
            # Use the first amount (pesos column)
            monto = _parse_ar_money(amounts[0])
            moneda = "ARS"

        if monto is None or monto == 0:
            continue

        tipo = "Ingreso" if monto < 0 else "Gasto"

        transactions.append({
            "fecha": f"{year}-{month}-{day}",
            "descripcion": desc,
            "monto": round(abs(monto), 2),
            "moneda": moneda,
            "tipo": tipo,
            "categoria_id": None,
            "subcategoria_id": None,
        })

    # Statement period
    statement_period = ""
    if transactions:
        from collections import Counter
        months = [t["fecha"][0:7] for t in transactions]
        if months:
            statement_period = Counter(months).most_common(1)[0][0]

    return {
        "transactions": transactions,
        "statement_period": statement_period,
        "card_or_account": card_info,
    }


def _parse_bbva_caja_ahorro(full_text: str) -> dict:
    """Parse BBVA Caja de Ahorro PDF text format into transactions."""
    lines = full_text.split("\n")

    transactions: list[dict] = []
    statement_year = ""

    # Extract year from document — look for "Vencimiento DD.YY" or similar
    for line in lines:
        m = re.search(r"Vencimiento\s+\d{1,2}\.(\d{2})\b", line)
        if m:
            statement_year = f"20{m.group(1)}"
            break
    if not statement_year:
        # Fallback: look for "Período DD/MM/YYYY" or 4-digit year anywhere
        for line in lines:
            m = re.search(r"(\d{4})", line)
            if m and m.group(1).startswith("20"):
                statement_year = m.group(1)
                break
    if not statement_year:
        statement_year = "2026"

    # Skip patterns — non-transaction entries
    skip_patterns = [
        "SALDO ANTERIOR", "TOTAL MOVIMIENTOS", "SALDO AL",
        "INTERESES GANADOS", "IVA TASA GENERAL", "COMISION POR RENTA",
        "IMPUESTO SOBRE", "COMISION", "IMPUESTO",
    ]

    # Section boundaries: only parse "Movimientos en cuentas"
    in_movimientos_section = False
    stop_sections = [
        "Transferencias", "Inversiones", "Legales y avisos",
        "OPERACIONES REALIZADAS", "LIQUIDACIONES DE DERECHOS",
    ]

    # Parse each line looking for DD/MM pattern at start
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Detect start of "Movimientos en cuentas" section
        if "Movimientos en cuentas" in stripped:
            in_movimientos_section = True
            continue

        # Detect end of section
        if in_movimientos_section and any(s in stripped for s in stop_sections):
            in_movimientos_section = False
            continue

        if not in_movimientos_section:
            continue

        # Skip column headers
        if "FECHA" in stripped and "CONCEPTO" in stripped:
            continue

        # Skip non-transaction rows
        if any(skip.lower() in stripped.lower() for skip in skip_patterns):
            continue

        # Try to match transaction line
        # Format: DD/MM  NNN  DESCRIPTION  [DEBITO]  [CREDITO]  SALDO
        # We need a flexible approach since columns may or may not have values
        m = re.match(r"^\s*(\d{2}/\d{2})\s+(\d{2,3})\s+(.+)$", stripped)
        if not m:
            continue

        fecha_str = m.group(1)  # DD/MM
        concepto_and_amounts = m.group(3).strip()

        # Extract amounts from the end of the line
        # Amounts look like: -749.993,85 or 44.208,00 or -1.351.330,62
        # There can be 2 or 3 amounts at the end (debito, credito, saldo)
        amount_pattern = r"-?[\d]+(?:\.[\d]{3})*,\d{2}"
        amounts = re.findall(amount_pattern, concepto_and_amounts)

        if len(amounts) < 2:
            # Need at least debito/credito + saldo
            continue

        # The last amount is always SALDO — ignore it
        # If there are 3 amounts: debito, credito, saldo
        # If there are 2 amounts: either (debito, saldo) or (credito, saldo)
        # Determine by sign: negative = debito

        # Remove amounts from description
        desc = concepto_and_amounts
        for amt in amounts:
            desc = desc.replace(amt, "", 1)
        desc = re.sub(r"\s+", " ", desc).strip()

        if not desc or len(desc) < 2:
            continue

        # Parse the transaction amount (not saldo)
        # With 3 amounts: amounts[0]=debito, amounts[1]=credito, amounts[2]=saldo
        # With 2 amounts: amounts[0]=debito or credito, amounts[1]=saldo
        saldo_str = amounts[-1]  # last is always saldo

        monto_val = None
        tipo = ""

        if len(amounts) >= 3:
            # Both debito and credito columns present
            debito_str = amounts[-3]
            credito_str = amounts[-2]
            debito_val = _parse_ar_money(debito_str)
            credito_val = _parse_ar_money(credito_str)
            if debito_val is not None and debito_val < 0:
                monto_val = abs(debito_val)
                tipo = "Gasto"
            elif credito_val is not None and credito_val > 0:
                monto_val = credito_val
                tipo = "Ingreso"
        else:
            # Only one amount + saldo
            val = _parse_ar_money(amounts[0])
            if val is not None:
                if val < 0:
                    monto_val = abs(val)
                    tipo = "Gasto"
                else:
                    monto_val = val
                    tipo = "Ingreso"

        if monto_val is None or monto_val == 0:
            continue

        # Build full date
        day, month = fecha_str.split("/")
        fecha = f"{statement_year}-{month}-{day}"

        transactions.append({
            "fecha": fecha,
            "descripcion": desc,
            "monto": round(monto_val, 2),
            "moneda": "ARS",
            "tipo": tipo,
            "categoria_id": None,
            "subcategoria_id": None,
        })

    # Determine statement period from transactions
    statement_period = ""
    if transactions:
        # Use the most common month among transactions
        from collections import Counter
        months = [t["fecha"][0:7] for t in transactions]
        if months:
            statement_period = Counter(months).most_common(1)[0][0]

    return {
        "transactions": transactions,
        "statement_period": statement_period,
        "card_or_account": "BBVA Caja de Ahorro",
    }


def _extract_pdf_text(pdf_bytes: bytes) -> str:
    """Extract text from PDF, trying PyPDF2 first, then pdfplumber as fallback."""
    import io

    # Try PyPDF2 first (faster)
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n".join(pages)
        if len(text.strip()) >= 100:
            return text
    except Exception:
        pass

    # Fallback to pdfplumber (handles more PDF formats)
    try:
        import pdfplumber
        pdf = pdfplumber.open(io.BytesIO(pdf_bytes))
        pages = [page.extract_text() or "" for page in pdf.pages]
        pdf.close()
        text = "\n".join(pages)
        if len(text.strip()) >= 100:
            return text
    except Exception:
        pass

    return ""


async def extract_from_pdf(pdf_bytes: bytes, categories: list[dict]) -> dict:
    """Extract transactions from a PDF. Detects Santander/BBVA format, falls back to AI."""
    full_text = _extract_pdf_text(pdf_bytes)

    if not full_text:
        return {
            "transactions": [],
            "error": "El PDF parece ser escaneado. Probá subiendo una foto o screenshot del extracto.",
        }

    # Detect Santander format
    is_santander = "Santander" in full_text and ("VISA" in full_text or "RESUMEN DE CUENTA" in full_text)

    if is_santander:
        result = _parse_santander_pdf_text(full_text)
        if result["transactions"] and categories:
            result["transactions"] = await _ai_categorize_batch(result["transactions"], categories)
        return result

    # Detect BBVA format
    is_bbva = "BBVA" in full_text or "AVBB" in full_text  # "AVBB" is reversed "BBVA" in some PDF encodings

    if is_bbva:
        result = _parse_bbva_pdf_text(full_text)
        if result["transactions"] and categories:
            result["transactions"] = await _ai_categorize_batch(result["transactions"], categories)
        return result

    # Generic: send to AI
    return await extract_from_text(full_text, categories)


def _parse_ar_money(val: str | None) -> float | None:
    """Parse Argentine money format: $1.219.066,31 → 1219066.31, U$S698,57 → 698.57"""
    if not val:
        return None
    s = str(val).strip()
    if not s or s == "None":
        return None
    # Remove currency symbols and spaces (order matters: U$S before $)
    for sym in ("U$S", "US$", "u$s", "us$", "$"):
        s = s.replace(sym, "")
    s = s.strip()
    if not s:
        return None
    # Handle negative: $-5.795.296,50
    negative = False
    if s.startswith("-"):
        negative = True
        s = s[1:]
    # Argentine format: dots are thousands, comma is decimal
    s = s.replace(".", "").replace(",", ".")
    try:
        val_f = float(s)
        return -val_f if negative else val_f
    except ValueError:
        return None


def _is_header_row(row_values: list) -> bool:
    """Check if a row is a header row (Fecha | Descripción | ...)."""
    vals = [str(v or "").strip().lower() for v in row_values]
    return "fecha" in vals and "descripción" in vals


def _is_section_header(row_values: list) -> str | None:
    """Detect section headers like 'Pago de tarjeta' or 'Otros conceptos'."""
    first = str(row_values[0] or "").strip().lower()
    if "pago de tarjeta" in first:
        return "pagos"
    if "otros conceptos" in first:
        return "otros"
    if "tarjeta de" in first or "tarjetas incluidas" in first:
        return "tarjeta_header"
    if "total de" in first:
        return "total"
    return None


async def extract_from_excel(excel_bytes: bytes, categories: list[dict]) -> dict:
    """Extract transactions from an Excel bank statement (Santander format)."""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]

    transactions = []
    current_section = None
    last_date = None
    card_info = ""
    statement_period = ""

    # Detect card info from sheet name
    card_info = ws.title or ""

    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if not any(v for v in vals):
            continue

        first = str(vals[0] or "").strip()

        # Detect sections
        section = _is_section_header(vals)
        if section:
            current_section = section
            continue

        # Skip header rows
        if _is_header_row(vals):
            continue

        # Skip total rows
        if first.lower().startswith("total de"):
            continue

        # Detect statement period from "Fecha de cierre"
        if first.lower() == "fecha de cierre":
            continue
        if not statement_period and len(vals) >= 2:
            cierre = str(vals[0] or "")
            if "/" in cierre and len(cierre) == 10:
                # Could be a date like 26/03/2026
                try:
                    parts = cierre.split("/")
                    if len(parts) == 3 and len(parts[2]) == 4:
                        statement_period = f"{parts[2]}-{parts[1]}"
                except Exception:
                    pass

        # Skip non-transaction sections
        if current_section in ("otros", "total"):
            continue

        # Parse transaction rows (6 columns: Fecha, Descripción, Cuotas, Comprobante, Monto pesos, Monto dólares)
        if len(vals) < 5:
            continue

        fecha_raw = vals[0]
        descripcion = str(vals[1] or "").strip()
        cuotas = str(vals[2] or "").strip()
        monto_pesos = vals[4] if len(vals) > 4 else None
        monto_dolares = vals[5] if len(vals) > 5 else None

        if not descripcion:
            continue

        # Skip payment/refund entries and administrative entries
        desc_lower = descripcion.lower()
        if any(skip in desc_lower for skip in [
            "su pago en pesos", "transferencia deuda", "aviso importante",
            "movimientos del resumen", "cierres y vencimientos",
        ]):
            continue

        # Handle date (may be empty = same as previous)
        if fecha_raw:
            fecha_str = str(fecha_raw).strip()
            if "/" in fecha_str:
                try:
                    parts = fecha_str.split("/")
                    if len(parts) == 3:
                        d, m, y = parts
                        if len(y) == 4:
                            last_date = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
                except Exception:
                    pass

        if not last_date:
            continue

        # Parse amount
        monto_pesos_val = _parse_ar_money(str(monto_pesos)) if monto_pesos else None
        monto_dolares_val = _parse_ar_money(str(monto_dolares)) if monto_dolares else None

        if monto_pesos_val is not None and monto_pesos_val != 0:
            monto = abs(monto_pesos_val)
            moneda = "ARS"
            tipo = "Ingreso" if monto_pesos_val < 0 else "Gasto"
        elif monto_dolares_val is not None and monto_dolares_val != 0:
            monto = abs(monto_dolares_val)
            moneda = "USD"
            tipo = "Ingreso" if monto_dolares_val < 0 else "Gasto"
        else:
            continue

        # Clean description
        desc_clean = descripcion.strip()
        if cuotas:
            desc_clean = f"{desc_clean} ({cuotas})"

        transactions.append({
            "fecha": last_date,
            "descripcion": desc_clean,
            "monto": round(monto, 2),
            "moneda": moneda,
            "tipo": tipo,
            "categoria_id": None,
            "subcategoria_id": None,
        })

    wb.close()

    # Use AI to suggest categories if we have categories
    if transactions and categories:
        transactions = await _ai_categorize_batch(transactions, categories)

    return {
        "transactions": transactions,
        "statement_period": statement_period,
        "card_or_account": card_info,
    }


async def _ai_categorize_chunk(
    transactions: list[dict],
    categories: list[dict],
    offset: int,
) -> list[dict]:
    """Categorize a chunk of transactions with AI. Returns suggestions with original indices."""
    cats_json = _build_categories_json(categories)
    descs_text = "\n".join(f"{i+1}. {t['descripcion']}" for i, t in enumerate(transactions))

    prompt = f"""Sos un experto en identificar comercios argentinos de resúmenes de tarjeta de crédito.
Para cada transacción, identificá qué comercio es y sugerí la categoría.

Transacciones:
{descs_text}

Categorías disponibles:
{cats_json}

Para cada una respondé: index (1-based), comercio_identificado, categoria_id (int o null), subcategoria_id (int o null), confianza ("alta"/"media"/"baja").
Respondé SOLO JSON: {{"suggestions": [...]}}"""

    try:
        client = _get_client()
        response = await client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            max_completion_tokens=2000,
        )
        raw = response.choices[0].message.content or "{}"
        result = json.loads(raw)
        return result.get("suggestions", [])
    except Exception as e:
        logger.error("AI chunk categorization failed (offset %d): %s", offset, e)
        return []


async def _ai_categorize_batch(transactions: list[dict], categories: list[dict]) -> list[dict]:
    """Categorize transactions in parallel chunks for speed."""
    import asyncio

    CHUNK_SIZE = 12
    chunks = []
    for i in range(0, len(transactions), CHUNK_SIZE):
        chunk = transactions[i:i + CHUNK_SIZE]
        chunks.append((chunk, i))

    # Run all chunks in parallel
    tasks = [
        _ai_categorize_chunk(chunk, categories, offset)
        for chunk, offset in chunks
    ]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    # Apply results back to transactions
    for (chunk, offset), result in zip(chunks, results):
        if isinstance(result, Exception):
            logger.error("Chunk at offset %d failed: %s", offset, result)
            continue
        for s in result:
            idx = s.get("index", 0) - 1  # 1-based to 0-based within chunk
            global_idx = offset + idx
            if 0 <= global_idx < len(transactions):
                transactions[global_idx]["categoria_id"] = s.get("categoria_id")
                transactions[global_idx]["subcategoria_id"] = s.get("subcategoria_id")
                transactions[global_idx]["confianza"] = s.get("confianza", "baja")
                if s.get("comercio_identificado"):
                    transactions[global_idx]["comercio_identificado"] = s["comercio_identificado"]

    return transactions


async def parse_statement(
    file_bytes: bytes,
    filename: str,
    mime_type: str,
    categories: list[dict],
) -> dict:
    """Dispatch to the right extractor based on file type."""
    fname_lower = filename.lower()
    if mime_type in ("image/jpeg", "image/jpg", "image/png", "image/webp"):
        return await extract_from_image(file_bytes, mime_type, categories)
    elif fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls") or \
         mime_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       "application/vnd.ms-excel"):
        return await extract_from_excel(file_bytes, categories)
    elif mime_type == "application/pdf" or fname_lower.endswith(".pdf"):
        return await extract_from_pdf(file_bytes, categories)
    else:
        return {"transactions": [], "error": f"Formato no soportado: {mime_type}"}

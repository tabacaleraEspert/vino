"""Generate downloadable Excel template for bank statement upload."""
from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_template() -> bytes:
    """Generate an Excel template for manual statement upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    example_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    # Headers
    headers = ["Fecha", "Descripcion", "Monto", "Moneda", "Tipo", "Cuotas"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Example rows
    examples = [
        ("15/04/2026", "Carrefour Express", 18420, "ARS", "Gasto", ""),
        ("14/04/2026", "Netflix", 6990, "ARS", "Gasto", ""),
        ("13/04/2026", "YPF Av. Libertador", 28500, "ARS", "Gasto", ""),
        ("12/04/2026", "Nike Alcorta", 45000, "ARS", "Gasto", "03/06"),
        ("10/04/2026", "Transferencia recibida", 150000, "ARS", "Ingreso", ""),
        ("08/04/2026", "Spotify", 4.99, "USD", "Gasto", ""),
    ]

    for row_idx, ex in enumerate(examples, 2):
        for col_idx, val in enumerate(ex, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = example_fill
            cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10

    # Instructions sheet
    ws2 = wb.create_sheet("Instrucciones")
    instructions = [
        ("Como usar esta plantilla", ""),
        ("", ""),
        ("1. Borra las filas de ejemplo en la hoja 'Movimientos'", ""),
        ("2. Completa tus movimientos con el formato indicado", ""),
        ("3. Subi el archivo en la app", ""),
        ("", ""),
        ("Columnas:", ""),
        ("Fecha", "DD/MM/YYYY (ej: 15/04/2026)"),
        ("Descripcion", "Nombre del comercio o descripcion del movimiento"),
        ("Monto", "Numero positivo, sin $ ni puntos de miles (ej: 18420)"),
        ("Moneda", "ARS o USD (si no pones nada, se asume ARS)"),
        ("Tipo", "Gasto o Ingreso (si no pones nada, se asume Gasto)"),
        ("Cuotas", "Opcional. Formato: cuota/total (ej: 03/06)"),
        ("", ""),
        ("Bancos con parser automatico:", ""),
        ("Santander", "PDF de resumen Visa y Excel de resumen"),
        ("Otros bancos", "Subi el PDF o screenshot y la IA lo extrae"),
    ]
    for row_idx, (a, b) in enumerate(instructions, 1):
        ws2.cell(row=row_idx, column=1, value=a).font = Font(bold=bool(a and not b))
        ws2.cell(row=row_idx, column=2, value=b)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
